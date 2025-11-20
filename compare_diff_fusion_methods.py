# -*- coding: utf-8 -*-
# @Author  : XinZhe Xie
# @University  : ZheJiang University
# This script compares different image fusion methods by calculating SSIM and PSNR metrics.
#
# Usage example:
# python compare_diff_fusion_methods.py \
#     --base_path /path/to/diff_fusion_methods_results \
#     --ground_truth_path /path/to/ground/truth \
#     --methods Method1 Method2 Method3 \
#     --datasets Dataset1 Dataset2 \
#     --output_dir ./evaluation_outputs

import numpy as np
import os
import pandas as pd
import cv2
import glob
from tqdm import tqdm
from skimage.metrics import structural_similarity as compare_ssim
from skimage.metrics import peak_signal_noise_ratio as compare_psnr
import warnings
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

warnings.filterwarnings('ignore')

def parse_args():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='Compare different image fusion methods by calculating SSIM and PSNR metrics')
    parser.add_argument('--base_path', type=str, default=r'E:\StackMFF Series Benchmark\results_of_different_methods',
                        help='Base path containing the result images for different methods')
    parser.add_argument('--ground_truth_path', type=str, default=r'E:\StackMFF Series Benchmark\results_of_different_methods\Ground Truth',
                        help='Path to ground truth images')
    parser.add_argument('--methods', type=str, nargs='+', default=['StackMFF V4'],
                        help='List of methods to compare')
    parser.add_argument('--datasets', type=str, nargs='+', 
                        default=['Mobile Depth','Middlebury','FlyingThings3D','Road-MF'],
                        help='List of datasets to evaluate')
    parser.add_argument('--enable_registration', action='store_true',
                        help='Enable image registration')
    parser.add_argument('--output_dir', type=str, default='./evaluation_outputs',
                        help='Directory to save output results')
    return parser.parse_args()

def create_output_dir(output_dir):
    """Create output directory if it doesn't exist"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

def get_image_formats(folder):
    """Get all image file formats in the folder"""
    formats = []
    for root, dirs, files in os.walk(folder):
        for file in files:
            filename, ext = os.path.splitext(file)
            ext = ext[1:].lower() # remove .
            if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'tiff']:
                if ext not in formats:
                    formats.append(ext)
    return formats

def align_images(img_result, img_truth):
    """
    Align images to make img_result aligned with img_truth
    Using ORB feature detection and matching, followed by affine transformation
    
    Args:
        img_result: Image to be registered
        img_truth: Reference image (Ground Truth)
        
    Returns:
        aligned_img: Registered image
    """
    # Convert to color image (if needed)
    if len(img_result.shape) == 2:
        img_result_color = cv2.cvtColor(img_result, cv2.COLOR_GRAY2BGR)
    else:
        img_result_color = img_result
        
    if len(img_truth.shape) == 2:
        img_truth_color = cv2.cvtColor(img_truth, cv2.COLOR_GRAY2BGR)
    else:
        img_truth_color = img_truth
    
    # Initialize ORB detector
    orb = cv2.ORB_create()
    
    # Find keypoints and descriptors
    kp1, des1 = orb.detectAndCompute(img_result_color, None)
    kp2, des2 = orb.detectAndCompute(img_truth_color, None)
    
    # If not enough keypoints are found, return the original image
    if des1 is None or des2 is None or len(des1) < 2 or len(des2) < 2:
        return img_result
    
    # Create BFMatcher object
    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
    
    # Match descriptors
    matches = bf.match(des1, des2)
    
    # If too few matching points, return the original image
    if len(matches) < 4:
        return img_result
    
    # Sort by distance
    matches = sorted(matches, key=lambda x: x.distance)
    
    # Extract coordinates of matching points
    src_pts = np.float32([kp1[m.queryIdx].pt for m in matches]).reshape(-1, 1, 2)
    dst_pts = np.float32([kp2[m.trainIdx].pt for m in matches]).reshape(-1, 1, 2)
    
    # Calculate homography matrix
    M, mask = cv2.findHomography(src_pts, dst_pts, cv2.RANSAC, 5.0)
    
    # If transformation matrix is found, perform perspective transformation
    if M is not None:
        aligned_img = cv2.warpPerspective(img_result, M, (img_truth.shape[1], img_truth.shape[0]))
        return aligned_img
    else:
        return img_result

def read_image_flexible(base_path, supported_formats=['jpg', 'jpeg', 'png']):
    """
    Read images more flexibly by trying multiple possible file extensions
    
    Args:
        base_path: File path without extension
        supported_formats: List of supported formats
    
    Returns:
        Image data or None
    """
    # First try the original path
    if os.path.exists(base_path):
        img = cv2.imread(base_path, 0)
        if img is not None:
            return img
    
    # Try various supported extensions
    for fmt in supported_formats:
        path_with_ext = f"{base_path}.{fmt}"
        if os.path.exists(path_with_ext):
            img = cv2.imread(path_with_ext, 0)
            if img is not None:
                return img
    
    # Try uppercase extensions
    for fmt in supported_formats:
        path_with_ext = f"{base_path}.{fmt.upper()}"
        if os.path.exists(path_with_ext):
            img = cv2.imread(path_with_ext, 0)
            if img is not None:
                return img
                
    return None

def get_images_dict(folder, supported_formats=['jpg', 'jpeg', 'png']):
    """
    Get all image files with supported formats in the folder, returning a dictionary with filenames (without extensions) as keys
    
    Args:
        folder: Folder path
        supported_formats: List of supported formats
    
    Returns:
        dict: {filename_without_ext: full_path}
    """
    images_dict = {}
    
    # Iterate through all supported formats
    for fmt in supported_formats:
        # Find files with lowercase extensions
        pattern = os.path.join(folder, f"*.{fmt}")
        for file_path in glob.glob(pattern):
            filename = os.path.basename(file_path)
            name_without_ext = os.path.splitext(filename)[0]
            images_dict[name_without_ext] = file_path
            
        # Find files with uppercase extensions
        pattern = os.path.join(folder, f"*.{fmt.upper()}")
        for file_path in glob.glob(pattern):
            filename = os.path.basename(file_path)
            name_without_ext = os.path.splitext(filename)[0]
            images_dict[name_without_ext] = file_path
    
    return images_dict

def main():
    # Parse command line arguments
    args = parse_args()
    
    # Create output directory
    create_output_dir(args.output_dir)
    
    # Define metrics (SSIM and PSNR only)
    metrics = ['SSIM', 'PSNR']
    
    # Create a dictionary to store results for all datasets
    all_datasets_results = {}
    # Dictionary to store method results across all datasets: {method: {dataset: {metric: value}}}
    all_methods_results = {method: {} for method in args.methods}
    
    # Get ground truth image paths
    for dataset in args.datasets:
        print('Testing:', dataset)
        
        # Get ground truth image dictionary {filename: full_path}
        gt_images_dict = get_images_dict(os.path.join(args.ground_truth_path, dataset))
        ground_truth_names = list(gt_images_dict.keys())
        
        if not ground_truth_names:
            print(f"Warning: No ground truth images found in {os.path.join(args.ground_truth_path, dataset)}")
            continue
            
        print(f"Found {len(ground_truth_names)} ground truth images")
        
        method_results = {}
        df = pd.DataFrame(columns=(['Method'] + metrics))
        
        # Iterate through methods
        for method_index, method in enumerate(args.methods):
            # Get image dictionary for this method's results
            method_images_dict = get_images_dict(os.path.join(args.base_path, method, dataset))
            
            if not method_images_dict:
                print(f"Warning: No result images found for method {method} in dataset {dataset}")
                continue
                
            print(f"Found {len(method_images_dict)} result images for method {method}")
            
            # Initialize metric dictionary
            metric_dict = {}
            for metric in metrics:
                metric_dict[metric] = 0  # Initialize to 0
            
            # Record number of successfully processed images
            valid_image_count = 0
            
            # Iterate through ground truth images
            for img_name in tqdm(ground_truth_names):
                # Get ground truth image path
                gt_img_path = gt_images_dict[img_name]
                img_truth = cv2.imread(gt_img_path, 0)
                
                if img_truth is None:
                    print(f"Warning: Could not read ground truth image {gt_img_path}")
                    continue
                
                # Construct base path for result image (without extension)
                result_base_path = os.path.join(args.base_path, method, dataset, img_name)
                
                # Try to read result image
                img_result = read_image_flexible(result_base_path)
                
                # Skip if result image cannot be read
                if img_result is None:
                    print(f"Warning: Could not read result image for {img_name} in method {method}")
                    continue
                
                # If registration is enabled and image sizes don't match, perform registration
                if args.enable_registration and img_result.shape != img_truth.shape:
                    img_result = align_images(img_result, img_truth)
                
                # Skip if image sizes still don't match
                if img_result.shape != img_truth.shape:
                    print(f"Warning: Image sizes still don't match after alignment for {img_name} in method {method}")
                    continue
                
                # Increase valid image count
                valid_image_count += 1
                
                # Calculate all metrics
                for metric_index, metric in enumerate(metrics):
                    try:
                        if metric == 'SSIM':
                            value = compare_ssim(img_truth, img_result, multichannel=False)
                        elif metric == 'PSNR':
                            value = compare_psnr(img_truth, img_result)
                        
                        metric_dict[metric] += value
                    except Exception as e:
                        print(f"Error calculating {metric} for {img_name}: {str(e)}")
                        # Set value to 0 on error
                        metric_dict[metric] += 0

            # Calculate averages using the actual number of successfully processed images
            if valid_image_count > 0:
                for metric in metric_dict:
                    metric_dict[metric] /= valid_image_count
                    metric_dict[metric] = round(metric_dict[metric], 4)
            else:
                print(f"Warning: No valid images processed for method {method}")

            method_results[method] = metric_dict
            # Store in all_methods_results
            all_methods_results[method][dataset] = metric_dict

        # Output results for current dataset
        for key, value in method_results.items():
            print(key, ':', value)
            temp = {'Method': key}
            new_row = {**temp, **value}

            # Create a new DataFrame with the new row
            new_df = pd.DataFrame([new_row])

            # Concatenate the new DataFrame with the existing one
            df = pd.concat([df, new_df], ignore_index=True)

            print(df)
        
        # Save results to Excel file
        output_filename = f'compare_result_{dataset}.xlsx'
        output_filepath = os.path.join(args.output_dir, output_filename)
        df.to_excel(output_filepath, index=False)
        print(f"Results saved to {output_filepath}")
        
        # Store results for current dataset
        all_datasets_results[dataset] = df
    
    # After all datasets are processed, create merged table and save
    if len(args.datasets) > 1 and all_methods_results:
        # Build data rows
        data_rows = []
        for method in args.methods:
            row = [method]
            for dataset in args.datasets:
                if dataset in all_methods_results[method]:
                    for metric in metrics:
                        value = all_methods_results[method][dataset].get(metric, '')
                        row.append(value)
                else:
                    # If no data for this dataset, fill with empty values
                    row.extend([''] * len(metrics))
            data_rows.append(row)
        
        # Save merged results to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'compare_result_merged_{timestamp}.xlsx'
        output_filepath = os.path.join(args.output_dir, output_filename)
        
        # Manually create Excel file and build multi-level header
        wb = Workbook()
        ws = wb.active
        ws.title = 'Comparison Results'
        
        # Write first row (dataset names)
        ws.cell(row=1, column=1, value='Datasets')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=1).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        col_idx = 2  # Start from column B (column A is Method)
        for dataset in args.datasets:
            start_col = col_idx
            end_col = col_idx + len(metrics) - 1
            # Merge cells
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=dataset)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            col_idx = end_col + 1
        
        # Write second row (metric names)
        ws.cell(row=2, column=1, value='Method')
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=1).fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        col_idx = 2
        for dataset in args.datasets:
            for metric in metrics:
                cell = ws.cell(row=2, column=col_idx, value=metric)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                col_idx += 1
        
        # Write data (starting from row 3)
        for row_idx, row_data in enumerate(data_rows, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:  # Method column
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Value column
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Save workbook
        wb.save(output_filepath)
        
        print(f"\nMerged results saved to {output_filepath}")
    
    # Print summary results
    print("\n" + "=" * 60)
    print("FINAL SUMMARY OF ALL DATASETS")
    print("=" * 60)
    
    for dataset, df in all_datasets_results.items():
        print(f"\nDataset: {dataset}")
        print("-" * 40)
        print(df.to_string(index=False))
    
    print("\n" + "=" * 60)
    print("END OF EVALUATION")
    print("=" * 60)

if __name__ == '__main__':
    main()
