# -*- coding: utf-8 -*-
# @Author  : XinZhe Xie
# @University  : ZheJiang University
# This script compares the final results when different fusion models are used as conditional images 
# for the second stage input of GMFF.
#
# Usage example:
# python compare_diff_stage1_input.py \
#     --base_path /path/to/stage1_diff_fusion_methods_stage2_gmff \
#     --methods CVT DWT DCT DTCWT NSCT IFCNN-MAX U2Fusion SDNet MFF-GAN SwinFusion MUFusion SwinMFF DDBFusion CCSR-Net MCCSR-Net "Zerene Stacker - DMap" "Zerene Stacker - PMax" "Helicon Focus 8 - A" "Helicon Focus 8 - B" "Helicon Focus 8 - C" StackMFF "StackMFF V2" "StackMFF V3" "StackMFF V4" \
#     --datasets "Mobile Depth" Middlebury \
#     --metrics BRISQUE PIQE \
#     --output_dir ./evaluation_outputs

import numpy as np
import pandas as pd
import cv2
import glob
from tqdm import tqdm
import warnings
import argparse
import torch
import pyiqa
from datetime import datetime
import os
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'

warnings.filterwarnings('ignore')

def parse_args():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='Compare different image fusion methods by calculating evaluation metrics')
    parser.add_argument('--base_path', type=str, default=r'/media/user/68fdd01e-c642-4deb-9661-23b76592afb1/xxz/project_image_stack_fusion/GMFF/results_diff_input_methods',
                        help='Base path containing the result images for different methods')
    parser.add_argument('--methods', type=str, nargs='+', default=['StackMFF V4'],
                        help='List of methods to compare')
    # parser.add_argument('--methods', type=str, nargs='+', default=['CVT', 
    #                                                                 'DWT', 
    #                                                                 'DCT', 
    #                                                                 'DTCWT', 
    #                                                                 'NSCT', 
    #                                                                 'IFCNN-MAX', 
    #                                                                 'U2Fusion', 
    #                                                                 'SDNet', 
    #                                                                 'MFF-GAN', 
    #                                                                 'SwinFusion', 
    #                                                                 'MUFusion', 
    #                                                                 'SwinMFF', 
    #                                                                 'DDBFusion', 
    #                                                                 'CCSR-Net', 
    #                                                                 'MCCSR-Net', 
    #                                                                 'Zerene Stacker - DMap', 
    #                                                                 'Zerene Stacker - PMax', 
    #                                                                 'Helicon Focus 8 - A', 
    #                                                                 'Helicon Focus 8 - B', 
    #                                                                 'Helicon Focus 8 - C', 
    #                                                                 'StackMFF', 
    #                                                                 'StackMFF V2', 
    #                                                                 'StackMFF V3', 
    #                                                                 'StackMFF V4']
    #                                                                 ,
    #                     help='List of methods to compare')
    parser.add_argument('--datasets', type=str, nargs='+', 
                        default=['Mobile Depth','Middlebury'],
                        help='List of datasets to evaluate')
    parser.add_argument('--metrics', type=str, nargs='+', default=['BRISQUE', 'PIQE'],
                        help='List of metrics to calculate')
    parser.add_argument('--enable_registration', default='false',
                        help='Enable image registration')
    parser.add_argument('--output_dir', type=str, default='./evaluation_outputs',
                        help='Directory to save output results')
    
    return parser.parse_args()

# Initialize BRISQUE model (global variable to avoid repeated initialization)
brisque_model = None

# Initialize PIQE model (global variable to avoid repeated initialization)
piqe_model = None

def get_brisque_model():
    """Get or initialize BRISQUE model"""
    global brisque_model
    if brisque_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        brisque_model = pyiqa.create_metric('brisque', device=device)
    return brisque_model

def get_piqe_model():
    """Get or initialize PIQE model"""
    global piqe_model
    if piqe_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        piqe_model = pyiqa.create_metric('piqe', device=device)
    return piqe_model

def calculate_pyiqa_metric(img, metric_model):
    """Calculate no-reference image quality metric using pyiqa
    
    Args:
        img: numpy array, shape (H, W) or (H, W, 3), range [0, 255], BGR format
        metric_model: pyiqa metric model
    
    Returns:
        Metric value
    """
    # 转换为RGB格式
    if len(img.shape) == 2:
        img_rgb = cv2.cvtColor(img, cv2.COLOR_GRAY2RGB)
    else:
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    
    # 转换为float并归一化到[0, 1], 格式: [H, W, C]
    img_rgb = img_rgb.astype(np.float32) / 255.0
    
    # 转换为tensor [H,W,C] -> [C,H,W] -> [1,C,H,W]
    img_tensor = torch.from_numpy(img_rgb).permute(2, 0, 1).unsqueeze(0)
    
    # 如果有GPU，移动到GPU
    if torch.cuda.is_available():
        img_tensor = img_tensor.cuda()
    
    # 计算指标
    with torch.no_grad():
        score = metric_model(img_tensor)
    
    return score.item()


# Create output directory
def create_output_dir(output_dir):
    """Create output directory if it doesn't exist"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

def get_image_formats(folder):
    """Get all image file formats in the folder"""
    formats = []
    for root, dirs, files in os.walk(folder):
        for file in files:
            filename, ext = os.path.splitext(file)
            ext = ext[1:].lower() # remove .
            if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'tiff']:
                if ext not in formats:
                    formats.append(ext)
    return formats

def read_image_flexible(base_path, supported_formats=['jpg', 'jpeg', 'png']):
    """
    Read images more flexibly, trying multiple possible file extensions
    
    Args:
        base_path: File path without extension
        supported_formats: List of supported formats
    
    Returns:
        Image data or None
    """
    # First try the original path
    if os.path.exists(base_path):
        img = cv2.imread(base_path, 0)
        if img is not None:
            return img
    
    # Try various supported extensions
    for fmt in supported_formats:
        path_with_ext = f"{base_path}.{fmt}"
        if os.path.exists(path_with_ext):
            img = cv2.imread(path_with_ext, 0)
            if img is not None:
                return img
    
    # Try uppercase extensions
    for fmt in supported_formats:
        path_with_ext = f"{base_path}.{fmt.upper()}"
        if os.path.exists(path_with_ext):
            img = cv2.imread(path_with_ext, 0)
            if img is not None:
                return img
                
    return None

def get_images_dict(folder, supported_formats=['jpg', 'jpeg', 'png']):
    """
    Get all image files in supported formats from a folder, returning a dictionary with filename (without extension) as key
    
    Args:
        folder: Folder path
        supported_formats: List of supported formats
    
    Returns:
        dict: {filename_without_ext: full_path}
    """
    images_dict = {}
    
    # Iterate through all supported formats
    for fmt in supported_formats:
        # Find files with lowercase extensions
        pattern = os.path.join(folder, f"*.{fmt}")
        for file_path in glob.glob(pattern):
            filename = os.path.basename(file_path)
            name_without_ext = os.path.splitext(filename)[0]
            images_dict[name_without_ext] = file_path
            
        # Find files with uppercase extensions
        pattern = os.path.join(folder, f"*.{fmt.upper()}")
        for file_path in glob.glob(pattern):
            filename = os.path.basename(file_path)
            name_without_ext = os.path.splitext(filename)[0]
            images_dict[name_without_ext] = file_path
    
    return images_dict

def main():
    # Parse command line arguments
    args = parse_args()
    
    # Create output directory
    create_output_dir(args.output_dir)
    
    # Output metric evaluation direction information
    print("\n" + "=" * 60)
    print("METRICS EVALUATION DIRECTION")
    print("=" * 60)
    device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
    for metric in args.metrics:
        try:
            if metric in ['BRISQUE', 'PIQE']:
                metric_model = pyiqa.create_metric(metric.lower(), device=device)
                direction = "Lower is better" if metric_model.lower_better else "Higher is better"
                print(f"{metric:12s}: {direction}")
            else:
                print(f"{metric:12s}: Unknown metric (not in pyiqa)")
        except Exception as e:
            print(f"{metric:12s}: Error checking direction - {str(e)}")
    print("=" * 60 + "\n")
    
    # Create a dictionary to store results for all datasets
    all_datasets_results = {}
    # Dictionary to store method results across all datasets: {method: {dataset: {metric: value}}}
    all_methods_results = {method: {} for method in args.methods}
    
    # Process each dataset
    for dataset in args.datasets:
        print('Testing:', dataset)
        
        method_results = {}
        df = pd.DataFrame(columns=(['Method'] + args.metrics))
        
        # Iterate through methods
        for method_index, method in enumerate(args.methods):
            # Get image dictionary for this method's results
            method_images_dict = get_images_dict(os.path.join(args.base_path, method, dataset))
            
            if not method_images_dict:
                print(f"Warning: No result images found for method {method} in dataset {dataset}")
                continue
                
            print(f"Found {len(method_images_dict)} result images for method {method}")
            
            # Initialize metric dictionary
            metric_dict = {}
            for metric in args.metrics:
                metric_dict[metric] = 0  # Initialize to 0
            
            # Record number of successfully processed images
            valid_image_count = 0
            
            # Iterate through images
            for img_name in tqdm(method_images_dict.keys()):
                # Get image path
                img_path = method_images_dict[img_name]
                
                # Read image
                img = cv2.imread(img_path, 0)  # Read as grayscale
                
                if img is None:
                    print(f"Warning: Could not read image {img_path}")
                    continue
                
                # Increase valid image count
                valid_image_count += 1
                
                # Calculate all metrics
                for metric_index, metric in enumerate(args.metrics):
                    
                    try:
                        if metric == 'BRISQUE':
                            brisq = get_brisque_model()
                            value = calculate_pyiqa_metric(img, brisq)
                        elif metric == 'PIQE':
                            piqe = get_piqe_model()
                            value = calculate_pyiqa_metric(img, piqe)
                        else:
                            print(f"Warning: Unknown metric {metric}, skipping...")
                            continue
                        
                        metric_dict[metric] += value
                    except Exception as e:
                        print(f"Error calculating {metric} for {img_name}: {str(e)}")
                        # Set value to 0 on error
                        metric_dict[metric] += 0

            # Calculate averages using the actual number of successfully processed images
            if valid_image_count > 0:
                for metric in metric_dict:
                    metric_dict[metric] /= valid_image_count
                    metric_dict[metric] = round(metric_dict[metric], 4)
            else:
                print(f"Warning: No valid images processed for method {method}")

            method_results[method] = metric_dict
            # Store in all_methods_results
            all_methods_results[method][dataset] = metric_dict

        # Output results for current dataset
        for key, value in method_results.items():
            print(key, ':', value)
            temp = {'Method': key}
            new_row = {**temp, **value}

            # Create a new DataFrame with the new row
            new_df = pd.DataFrame([new_row])

            # Concatenate the new DataFrame with the existing one
            df = pd.concat([df, new_df], ignore_index=True)

            print(df)
        
        # Store results for current dataset
        all_datasets_results[dataset] = df
    
    # After all datasets are processed, create merged table and save
    if len(args.datasets) > 1 and all_methods_results:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        
        # Build data rows
        data_rows = []
        for method in args.methods:
            row = [method]
            for dataset in args.datasets:
                if dataset in all_methods_results[method]:
                    for metric in args.metrics:
                        value = all_methods_results[method][dataset].get(metric, '')
                        row.append(value)
                else:
                    # If no data for this dataset, fill with empty values
                    row.extend([''] * len(args.metrics))
            data_rows.append(row)
        
        # Save merged results to Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'compare_result_merged_{timestamp}.xlsx'
        output_filepath = os.path.join(args.output_dir, output_filename)
        
        # Manually create Excel file and build multi-level header
        wb = Workbook()
        ws = wb.active
        ws.title = 'Comparison Results'
        
        # Write first row (dataset names)
        ws.cell(row=1, column=1, value='Datasets')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=1).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        col_idx = 2  # Start from column B (column A is Method)
        for dataset in args.datasets:
            start_col = col_idx
            end_col = col_idx + len(args.metrics) - 1
            # Merge cells
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=dataset)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            col_idx = end_col + 1
        
        # Write second row (metric names)
        ws.cell(row=2, column=1, value='Method')
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=1).fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        col_idx = 2
        for dataset in args.datasets:
            for metric in args.metrics:
                # Add arrows based on metric type
                metric_name = f"{metric} ↓" if metric in ['BRISQUE', 'PIQE'] else f"{metric} ↑"
                cell = ws.cell(row=2, column=col_idx, value=metric_name)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                col_idx += 1
        
        # Write data (starting from row 3)
        for row_idx, row_data in enumerate(data_rows, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:  # Method column
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # Value column
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Save workbook
        wb.save(output_filepath)
        
        print(f"\nMerged results saved to {output_filepath}")
    
    # Print summary results
    print("\n" + "=" * 60)
    print("FINAL SUMMARY OF ALL DATASETS")
    print("=" * 60)
    
    for dataset, df in all_datasets_results.items():
        print(f"\nDataset: {dataset}")
        print("-" * 40)
        print(df.to_string(index=False))
    
    # Output metric evaluation direction reminder again
    print("\n" + "=" * 60)
    print("METRICS EVALUATION DIRECTION (REMINDER)")
    print("=" * 60)
    device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
    for metric in args.metrics:
        try:
            if metric in ['BRISQUE', 'PIQE']:
                metric_model = pyiqa.create_metric(metric.lower(), device=device)
                direction = "Lower is better ↓" if metric_model.lower_better else "Higher is better ↑"
                print(f"{metric:12s}: {direction}")
            else:
                print(f"{metric:12s}: Unknown metric (not in pyiqa)")
        except Exception as e:
            print(f"{metric:12s}: Error checking direction - {str(e)}")
    print("=" * 60)
    
    print("\n" + "=" * 60)
    print("END OF EVALUATION")
    print("=" * 60)

if __name__ == '__main__':
    main()